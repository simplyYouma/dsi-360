"""Modèle Excel d'import d'inventaire : un classeur prêt à remplir par le métier.

Trois feuilles :
- **Inventaire** — la grille de saisie. La ligne d'en-tête reprend *exactement* les intitulés que
  l'import sait lire ; en dessous, tout est vide, prêt à remplir. Chaque en-tête porte une note
  (survol) courte ; le détail complet est dans le mode d'emploi.
- **Mode d'emploi** — mis en page : bandeau, tableau des colonnes (obligatoire ? / ce qu'on y met),
  et la règle de création / mise à jour.
- **Listes** — masquée : elle alimente les listes déroulantes. Ses valeurs sont celles **réellement
  présentes dans la plateforme** au moment du téléchargement, pas une liste figée dans le code.

Type, emplacement et département se choisissent dans une liste déroulante **sans être enfermés
dedans** : la validation ne refuse rien, on peut taper un nouveau site ou un nouveau type, qui sera
créé à l'import. Proposer sans imposer est le seul moyen d'éviter à la fois « AGENCE KAYES » /
« Agence 11 Kayes » / « Kayes » et le blocage sur un matériel dont le lieu n'existe pas encore.

On ne met **aucune ligne d'exemple** dans la feuille de saisie : elle serait importée comme un vrai
équipement. L'exemple vit dans le mode d'emploi, que l'import ignore (il ne lit que la feuille
portant les en-têtes).
"""

from collections.abc import Mapping, Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

# Charte, calquée sur les exports : en-tête noir, texte blanc. La colonne obligatoire vire au rouge.
_NOIR = "16181D"
_ROUGE = "C0392B"
_VERT_CLAIR = "EEF7DF"
_GRIS_LIGNE = "E6E8EC"
_GRIS_FOND = "F6F7F9"

# (en-tête EXACT — doit correspondre aux alias du lecteur —, obligatoire, ce qu'on y met).
_COLONNES: tuple[tuple[str, bool, str], ...] = (
    ("Code immo", True,
     "Code d'immobilisation comptable. C'est la CLÉ de reconnaissance : un code déjà connu met à "
     "jour la fiche, un code nouveau crée l'équipement. Sans lui, la ligne est ignorée — rien ne "
     "permettrait de la retrouver au prochain import."),
    ("Désignation", True,
     "Ce qu'est le matériel (ex. « Ordinateur portable Dell Latitude 5540 »). "
     "Sans elle, la ligne est ignorée."),
    ("Type", False,
     "Nature (Ordinateur portable, Serveur, Imprimante…). À choisir dans la liste déroulante, "
     "ou à taper librement : un type inconnu est créé à l'import."),
    ("N° série", False,
     "Numéro de série constructeur. Unique dans tout le parc : si deux lignes portent le même, "
     "ou s'il est déjà pris, il est écarté pour cette ligne et le compte-rendu le signale."),
    ("Modèle", False, "Référence du modèle (ex. Latitude 5540)."),
    ("Emplacement", False,
     "Où se trouve le matériel (agence, siège…). À choisir dans la liste déroulante, ou à taper "
     "librement : un emplacement inconnu est créé à l'import."),
    ("Département", False,
     "Service rattaché. À choisir dans la liste déroulante, ou à taper librement : un département "
     "inconnu est créé à l'import."),
    ("Matricule", False,
     "Matricule de l'agent détenteur. Rapproché d'un compte s'il existe, sinon conservé tel quel."),
    ("Taux", False, "Taux d'amortissement annuel, en % (ex. 25)."),
    ("Date acquisition", False, "Date d'achat (jj/mm/aaaa)."),
    ("Durée", False, "Durée d'amortissement, en années (ex. 4)."),
    ("Valeur acquisition", False, "Prix d'achat en FCFA (ex. 850000)."),
    ("État constaté", False, "Dernier contrôle : Bon, Rebut ou Cassé. Vide si non contrôlé."),
)

#: Largeur d'une colonne : ajustée à l'intitulé (comme l'export), la désignation un peu plus large
#: car elle porte du texte long. Fini les colonnes démesurées.
_LARGEUR_DESIGNATION = 30

#: Intitulé -> rang de la colonne, pour la retrouver sans compter à la main.
_RANG = {titre: i for i, (titre, _, _) in enumerate(_COLONNES, start=1)}

_TRAIT = Side(style="thin", color=_GRIS_LIGNE)
_BORDURE = Border(left=_TRAIT, right=_TRAIT, top=_TRAIT, bottom=_TRAIT)


def _largeur(titre: str) -> int:
    if titre == "Désignation":
        return _LARGEUR_DESIGNATION
    return min(max(len(titre) + 3, 11), 20)


def _entete_saisie(feuille: Worksheet) -> None:
    """Ligne d'en-tête façon export : fond noir, texte blanc ; la colonne obligatoire en rouge."""
    blanc_gras = Font(bold=True, color="FFFFFF", size=11)
    fond_noir = PatternFill("solid", fgColor=_NOIR)
    fond_rouge = PatternFill("solid", fgColor=_ROUGE)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (titre, obligatoire, aide) in enumerate(_COLONNES, start=1):
        cellule = feuille.cell(row=1, column=i, value=titre)
        cellule.font = blanc_gras
        cellule.fill = fond_rouge if obligatoire else fond_noir
        cellule.alignment = centre
        note = Comment(("OBLIGATOIRE. " if obligatoire else "") + aide, "DSI 360")
        note.width = 220
        note.height = 110
        cellule.comment = note
        feuille.column_dimensions[get_column_letter(i)].width = _largeur(titre)
    feuille.row_dimensions[1].height = 30
    feuille.freeze_panes = "A2"  # l'en-tête reste visible en défilant


def _dropdown_etat(feuille: Worksheet) -> None:
    """Liste déroulante Bon / Rebut / Cassé sur la colonne « État constaté »."""
    indices = [i for i, c in enumerate(_COLONNES, start=1) if c[0] == "État constaté"]
    if not indices:
        return
    lettre = get_column_letter(indices[0])
    validation = DataValidation(
        type="list", formula1='"Bon,Rebut,Cassé"', allow_blank=True,
        showErrorMessage=True, errorTitle="Valeur non prévue",
        error="Choisissez Bon, Rebut ou Cassé — ou laissez vide.",
    )
    feuille.add_data_validation(validation)
    validation.add(f"{lettre}2:{lettre}{_LIGNES_SAISIE}")


#: Colonne de saisie -> (clé du référentiel, colonne de la feuille « Listes »).
_REFERENTIELS: tuple[tuple[str, str, str], ...] = (
    ("Type", "types", "A"),
    ("Emplacement", "emplacements", "B"),
    ("Département", "departements", "C"),
)

#: Au-delà, un classeur d'inventaire n'est plus rempli à la main.
_LIGNES_SAISIE = 1000


def _feuille_listes(feuille: Worksheet, listes: Mapping[str, Sequence[str]]) -> None:
    """Les valeurs qui alimentent les déroulantes, une par colonne. Feuille masquée.

    Masquée, mais pas protégée : qui veut voir ce que la plateforme connaît déjà peut l'afficher.
    """
    for titre, cle, colonne in _REFERENTIELS:
        feuille[f"{colonne}1"] = titre
        feuille[f"{colonne}1"].font = Font(bold=True, color=_NOIR)
        feuille.column_dimensions[colonne].width = 28
        for i, valeur in enumerate(listes.get(cle, ()), start=2):
            feuille[f"{colonne}{i}"] = valeur
    feuille.sheet_state = "hidden"


def _deroulantes_referentiels(
    saisie: Worksheet, listes: Mapping[str, Sequence[str]], nom_listes: str
) -> None:
    """Une déroulante par référentiel, **ouverte** : elle propose, elle n'impose pas.

    ``showErrorMessage=False`` est le cœur du réglage : Excel affiche la flèche et la liste, mais
    laisse passer une valeur tapée à la main. Un import crée le type ou le site inconnu ; refuser
    la saisie bloquerait le remplissage sur un matériel dont le lieu n'est pas encore référencé.
    """
    for titre, cle, colonne in _REFERENTIELS:
        valeurs = listes.get(cle, ())
        # Pas de liste vide : une flèche sans rien derrière déroute plus qu'elle n'aide.
        if not valeurs:
            continue
        position = _RANG.get(titre)
        if position is None:
            continue
        lettre = get_column_letter(position)
        validation = DataValidation(
            type="list",
            formula1=f"'{nom_listes}'!${colonne}$2:${colonne}${len(valeurs) + 1}",
            allow_blank=True,
            showErrorMessage=False,
            promptTitle=titre,
            prompt="Choisissez dans la liste, ou tapez une nouvelle valeur : elle sera creee.",
            showInputMessage=True,
        )
        saisie.add_data_validation(validation)
        validation.add(f"{lettre}2:{lettre}{_LIGNES_SAISIE}")


def _bandeau(feuille: Worksheet, ligne: int, texte: str) -> None:
    """Titre de section : pleine largeur (A:C), fond vert clair, gras."""
    feuille.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=3)
    cellule = feuille.cell(row=ligne, column=1, value=texte)
    cellule.font = Font(bold=True, size=11, color=_NOIR)
    cellule.fill = PatternFill("solid", fgColor=_VERT_CLAIR)
    cellule.alignment = Alignment(vertical="center", indent=1)
    feuille.row_dimensions[ligne].height = 22


_REGLES = (
    "• Le « Code immo » est la clé. Code déjà connu → la ligne MET À JOUR la fiche ; "
    "code nouveau → elle CRÉE un équipement.",
    "• Une ligne SANS code immo est ignorée, et comptée comme telle dans le compte-rendu : sans "
    "clé, on ne saurait ni éviter le doublon, ni la retrouver au prochain import.",
    "• À la mise à jour, les colonnes comptables (désignation, taux, date, durée, valeur) sont "
    "reprises du fichier.",
    "• Les colonnes de terrain (type, n° série, modèle, emplacement, département, matricule) ne "
    "sont remplies QUE si elles étaient vides : une correction faite à l'écran n'est jamais "
    "écrasée.",
    "• Réimporter le même fichier ne crée pas de doublon.",
    "• Deux colonnes sont obligatoires : « Code immo » et « Désignation ». Tout le reste peut "
    "rester vide — une case vide n'est jamais bloquante, elle laisse simplement la fiche "
    "incomplète, à compléter plus tard à l'écran.",
    "• Type, emplacement et département proposent une liste, sans l'imposer : une valeur tapée à "
    "la main est acceptée et créée à l'import.",
    "• La plateforme attribue elle-même une référence « INV-… » à chaque équipement : elle n'est "
    "pas dans le fichier, ne la saisissez pas.",
)


def _mode_emploi(feuille: Worksheet) -> None:
    """Feuille d'explications mise en page. Ignorée par l'import (pas d'en-têtes reconnues)."""
    feuille.sheet_view.showGridLines = False
    feuille.column_dimensions["A"].width = 22
    feuille.column_dimensions["B"].width = 14
    feuille.column_dimensions["C"].width = 74

    # Titre, pleine largeur, fond noir.
    feuille.merge_cells("A1:C1")
    titre = feuille.cell(row=1, column=1, value="  Modèle d'import de l'inventaire — mode d'emploi")
    titre.font = Font(bold=True, size=15, color="FFFFFF")
    titre.fill = PatternFill("solid", fgColor=_NOIR)
    titre.alignment = Alignment(vertical="center")
    feuille.row_dimensions[1].height = 34

    # Section 1 : le tableau des colonnes.
    _bandeau(feuille, 3, "Les colonnes, une à une")
    for j, texte in enumerate(("Colonne", "Obligatoire", "Ce qu'on y met"), start=1):
        cellule = feuille.cell(row=4, column=j, value=texte)
        cellule.font = Font(bold=True, color="FFFFFF")
        cellule.fill = PatternFill("solid", fgColor=_NOIR)
        cellule.alignment = Alignment(vertical="center", indent=1)
        cellule.border = _BORDURE
    feuille.row_dimensions[4].height = 20

    ligne = 5
    for titre_col, obligatoire, aide in _COLONNES:
        feuille.cell(row=ligne, column=1, value=titre_col).font = Font(bold=True, color=_NOIR)
        marque = feuille.cell(row=ligne, column=2, value="Oui" if obligatoire else "—")
        marque.alignment = Alignment(horizontal="center")
        if obligatoire:
            marque.font = Font(bold=True, color=_ROUGE)
        feuille.cell(row=ligne, column=3, value=aide)
        for j in (1, 2, 3):
            c = feuille.cell(row=ligne, column=j)
            c.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
            c.border = _BORDURE
            if ligne % 2 == 1:
                c.fill = PatternFill("solid", fgColor=_GRIS_FOND)
        feuille.row_dimensions[ligne].height = 30
        ligne += 1

    # Section 2 : création vs mise à jour.
    ligne += 1
    _bandeau(feuille, ligne, "Création ou mise à jour")
    ligne += 1
    for texte in _REGLES:
        feuille.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=3)
        cellule = feuille.cell(row=ligne, column=1, value=texte)
        cellule.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        feuille.row_dimensions[ligne].height = 32
        ligne += 1

    # Section 3 : un exemple concret.
    ligne += 1
    _bandeau(feuille, ligne, "Exemple")
    ligne += 1
    feuille.merge_cells(start_row=ligne, start_column=1, end_row=ligne, end_column=3)
    exemple = feuille.cell(
        row=ligne, column=1,
        value="INV00042 · Ordinateur portable Dell Latitude 5540 · Type : Ordinateur portable · "
        "Matricule : MAT-4507 · Taux : 25 · Date : 12/03/2024 · Durée : 4 · Valeur : 850000 · "
        "État : Bon",
    )
    exemple.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    exemple.font = Font(italic=True, color="414751")
    feuille.row_dimensions[ligne].height = 30


def construire_modele(listes: Mapping[str, Sequence[str]] | None = None) -> bytes:
    """Le classeur modèle, en octets .xlsx.

    ``listes`` porte les valeurs à proposer, par clé de référentiel (``types``, ``emplacements``,
    ``departements``) : ce sont celles de la plateforme, lues au moment du téléchargement. Sans
    elles, le modèle reste utilisable — il propose seulement moins.
    """
    listes = listes or {}
    classeur = Workbook()
    saisie = classeur.active
    assert saisie is not None  # un classeur neuf a toujours sa feuille active
    saisie.title = "Inventaire"
    _entete_saisie(saisie)
    _dropdown_etat(saisie)
    _mode_emploi(classeur.create_sheet("Mode d'emploi"))
    nom_listes = "Listes"
    _feuille_listes(classeur.create_sheet(nom_listes), listes)
    _deroulantes_referentiels(saisie, listes, nom_listes)
    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
