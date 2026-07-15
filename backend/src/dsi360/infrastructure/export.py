"""Génération de fichiers d'export : CSV (stdlib) et Excel (openpyxl)."""

import csv
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def vers_csv(entetes: list[str], lignes: list[list[Any]]) -> bytes:
    tampon = io.StringIO()
    ecrivain = csv.writer(tampon, delimiter=";")
    ecrivain.writerow(entetes)
    ecrivain.writerows(lignes)
    # BOM UTF-8 pour qu'Excel ouvre correctement les accents.
    return tampon.getvalue().encode("utf-8-sig")


def vers_xlsx(entetes: list[str], lignes: list[list[Any]], titre: str) -> bytes:
    classeur = Workbook()
    feuille = classeur.active
    assert feuille is not None  # un classeur neuf a toujours une feuille active  # noqa: S101
    feuille.title = titre[:31]  # Excel limite le nom d'onglet à 31 caractères

    feuille.append(entetes)
    gras = Font(bold=True, color="FFFFFF")
    fond = PatternFill("solid", fgColor="16181D")
    for cellule in feuille[1]:
        cellule.font = gras
        cellule.fill = fond

    for ligne in lignes:
        feuille.append(ligne)

    for i, entete in enumerate(entetes, start=1):
        # get_column_letter gère au-delà de 26 colonnes (AA, AB…) ; chr(64+i) produisait des
        # lettres invalides. lignes ragged : on ne lit la cellule que si elle existe.
        largeurs = [len(entete)] + [len(str(lg[i - 1])) for lg in lignes if len(lg) >= i]
        feuille.column_dimensions[get_column_letter(i)].width = min(max(largeurs) + 2, 50)

    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()
