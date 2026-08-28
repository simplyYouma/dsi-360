"""Import de l'inventaire : charger sans jamais écraser le travail de la DSI.

C'est LE point délicat du module. Le fichier vient de la comptabilité, mais la DSI corrige à
l'écran ce que la compta ne connaît pas (emplacement réel, n° de série, détenteur). Un réimport
qui reprendrait tout annulerait ce travail à chaque fois.
"""

from datetime import date
from io import BytesIO
from typing import Any

import openpyxl
import pytest
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from dsi360.application.import_equipements import importer_classeur
from dsi360.infrastructure.repositories import equipement as repo
from tests.integration.conftest import creer_utilisateur

# En-têtes du fichier de référence — « Designtion » compris (faute de frappe du fichier source).
_ENTETES = [
    "CODE IMMO",
    "MATRICULE",
    "N° Série",
    "MODELE",
    "Designtion",
    "EMPLACEMENT",
    "DEPARTEMENT",
    "Taux",
    "DA",
    "Durée",
    "VA",
    "Etat bon",
    "Rebut",
    "Casse",
]


def _classeur(lignes: list[list[Any]], preambule: bool = True) -> bytes:
    """Classeur de test, avec le préambule que portent les exports réels."""
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    assert feuille is not None
    if preambule:
        feuille.append([date(2026, 7, 20)])
        feuille.append(["FICHIERS DES INVENTAIRES 2026"])
    feuille.append(_ENTETES)
    for ligne in lignes:
        feuille.append(ligne)
    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


def _ligne(**champs: Any) -> list[Any]:
    base: dict[str, Any] = {
        "CODE IMMO": "INF00208",
        "MATRICULE": None,
        "N° Série": None,
        "MODELE": None,
        "Designtion": "GAB Total Missabougou",
        "EMPLACEMENT": "GAB EXT",
        "DEPARTEMENT": "GAB SYAMA",
        "Taux": 25,
        "DA": date(2005, 7, 22),
        "Durée": "4 ans",
        "VA": 23074595,
        "Etat bon": None,
        "Rebut": None,
        "Casse": None,
    }
    base.update(champs)
    return [base[e] for e in _ENTETES]


async def _acteur(session: AsyncSession, email: str) -> dict[str, Any]:
    uid = await creer_utilisateur(session, email=email, profil="ADMIN")
    return {"id": uid, "email": email}


async def test_un_premier_import_cree_le_parc(session: AsyncSession) -> None:
    acteur = await _acteur(session, "admin.imp1@afgbank.ml")

    rapport = await importer_classeur(session, _classeur([_ligne()]), acteur)

    assert rapport == {
        "total": 1,
        "crees": 1,
        "mis_a_jour": 0,
        "ignores": 0,
        "detenteurs_non_rapproches": 0,
        "avec_etat_constate": 0,
        "constats_enregistres": 0,
        "series_en_double": 0,
    }
    e = await repo.par_code_immo(session, "INF00208")
    assert e is not None
    assert e["designation"] == "GAB Total Missabougou"
    assert float(e["valeur_acquisition"]) == 23074595
    assert e["date_acquisition"] == date(2005, 7, 22)
    assert e["duree_annees"] == 4, "« 4 ans » doit être lu comme un nombre d'années"
    assert e["source"] == "IMPORT_IMMO"
    assert e["emplacement"] == "GAB EXT", "le référentiel s'alimente tout seul"
    assert e["departement"] == "GAB SYAMA"


async def test_reimporter_ne_cree_pas_de_doublon(session: AsyncSession) -> None:
    """Idempotence par code d'immobilisation : le fichier se recharge sans risque."""
    acteur = await _acteur(session, "admin.imp2@afgbank.ml")
    fichier = _classeur([_ligne()])

    await importer_classeur(session, fichier, acteur)
    second = await importer_classeur(session, fichier, acteur)

    assert second["crees"] == 0
    assert second["mis_a_jour"] == 1
    total = await session.scalar(
        text("SELECT count(*) FROM core.equipement WHERE code_immo = 'INF00208'")
    )
    assert total == 1


class TestRegleDOr:
    """La comptabilité écrase ses colonnes ; la DSI garde les siennes."""

    async def test_une_correction_de_la_dsi_survit_au_reimport(
        self, session: AsyncSession
    ) -> None:
        acteur = await _acteur(session, "admin.imp3@afgbank.ml")
        depart = _classeur([_ligne(EMPLACEMENT="GAB EXT", **{"N° Série": "SN-ORIGINE"})])
        await importer_classeur(session, depart, acteur)
        cree = await repo.par_code_immo(session, "INF00208")
        assert cree is not None

        # La DSI constate sur le terrain que le matériel est ailleurs, et corrige.
        vrai_lieu = await repo.trouver_ou_creer_referentiel(
            session, "emplacements", "Agence Yirimadio"
        )
        await repo.maj(
            session,
            cree["id"],
            {"emplacement_id": vrai_lieu, "numero_serie": "SN-CORRIGE", "modele": "SelfServ 25"},
        )
        await session.commit()

        # Le fichier suivant redit l'ancien emplacement — et revalorise le bien.
        await importer_classeur(
            session,
            _classeur([_ligne(EMPLACEMENT="GAB EXT", VA=9_000_000, **{"N° Série": "SN-ORIGINE"})]),
            acteur,
        )

        apres = await repo.par_code_immo(session, "INF00208")
        assert apres is not None
        assert apres["emplacement"] == "Agence Yirimadio", "la correction terrain tient"
        assert apres["numero_serie"] == "SN-CORRIGE", "le n° de série saisi n'est pas écrasé"
        assert apres["modele"] == "SelfServ 25"
        assert float(apres["valeur_acquisition"]) == 9_000_000, "la compta, elle, fait foi"

    async def test_l_import_comble_les_trous(self, session: AsyncSession) -> None:
        """Ne pas écraser ne veut pas dire ne rien remplir : une case vide se complète."""
        acteur = await _acteur(session, "admin.imp4@afgbank.ml")
        await importer_classeur(session, _classeur([_ligne()]), acteur)

        await importer_classeur(
            session, _classeur([_ligne(MODELE="NCR SelfServ 25")]), acteur
        )

        apres = await repo.par_code_immo(session, "INF00208")
        assert apres is not None
        assert apres["modele"] == "NCR SelfServ 25"

    async def test_les_colonnes_comptables_sont_toujours_reprises(
        self, session: AsyncSession
    ) -> None:
        acteur = await _acteur(session, "admin.imp5@afgbank.ml")
        await importer_classeur(session, _classeur([_ligne(Taux=25)]), acteur)

        await importer_classeur(
            session, _classeur([_ligne(Taux=20, **{"Durée": "5 ans"})]), acteur
        )

        apres = await repo.par_code_immo(session, "INF00208")
        assert apres is not None
        assert float(apres["taux"]) == 20
        assert apres["duree_annees"] == 5


class TestDetenteur:
    async def test_le_matricule_rattache_a_un_compte(self, session: AsyncSession) -> None:
        acteur = await _acteur(session, "admin.imp6@afgbank.ml")
        agent = await creer_utilisateur(session, email="porteur.imp@afgbank.ml")
        await session.execute(
            text("UPDATE core.utilisateur SET matricule = 'M-4501' WHERE id = cast(:i as uuid)"),
            {"i": agent},
        )
        await session.commit()

        rapport = await importer_classeur(session, _classeur([_ligne(MATRICULE="m-4501")]), acteur)

        assert rapport["detenteurs_non_rapproches"] == 0
        e = await repo.par_code_immo(session, "INF00208")
        assert e is not None
        assert e["detenteur_id"] == agent, "rattaché malgré la casse différente"

    async def test_un_matricule_inconnu_est_conserve_et_signale(
        self, session: AsyncSession
    ) -> None:
        """On n'invente pas de compte : le matricule brut reste, un import futur rattachera."""
        acteur = await _acteur(session, "admin.imp7@afgbank.ml")

        rapport = await importer_classeur(
            session, _classeur([_ligne(MATRICULE="M-INCONNU")]), acteur
        )

        assert rapport["detenteurs_non_rapproches"] == 1
        e = await repo.par_code_immo(session, "INF00208")
        assert e is not None
        assert e["detenteur_id"] is None
        assert e["matricule_brut"] == "M-INCONNU"


def _classeur_type(lignes: list[dict[str, Any]]) -> bytes:
    """Classeur au format du modèle : en-tête « Type » et colonne unique « État constaté »."""
    entetes = ["Code immo", "Désignation", "Type", "État constaté"]
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    assert feuille is not None
    feuille.append(entetes)
    for ligne in lignes:
        feuille.append([ligne.get(e) for e in entetes])
    tampon = BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


class TestType:
    """Le type d'inventaire s'importe comme le reste du terrain : rempli si vide, jamais écrasé."""

    async def test_le_type_est_lu_et_cree_a_la_volee(self, session: AsyncSession) -> None:
        acteur = await _acteur(session, "admin.imptype1@afgbank.ml")

        rapport = await importer_classeur(
            session,
            _classeur_type([
                {
                    "Code immo": "INV00500",
                    "Désignation": "Portable de test",
                    "Type": "Ordinateur portable",
                }
            ]),
            acteur,
        )

        assert rapport["crees"] == 1
        e = await repo.par_code_immo(session, "INV00500")
        assert e is not None
        assert e["type"] == "Ordinateur portable", "le type se crée tout seul et se rattache"

    async def test_un_type_saisi_a_l_ecran_survit_au_reimport(self, session: AsyncSession) -> None:
        acteur = await _acteur(session, "admin.imptype2@afgbank.ml")
        await importer_classeur(
            session,
            _classeur_type([{"Code immo": "INV00501", "Désignation": "Poste"}]),
            acteur,
        )
        cree = await repo.par_code_immo(session, "INV00501")
        assert cree is not None
        # La DSI classe le matériel à l'écran.
        tid = await repo.trouver_ou_creer_referentiel(session, "types", "Serveur")
        await repo.maj(session, cree["id"], {"type_id": tid})
        await session.commit()

        # Le fichier suivant propose un autre type : il ne doit pas écraser la saisie.
        await importer_classeur(
            session,
            _classeur_type([
                {"Code immo": "INV00501", "Désignation": "Poste", "Type": "Onduleur"}
            ]),
            acteur,
        )
        apres = await repo.par_code_immo(session, "INV00501")
        assert apres is not None
        assert apres["type"] == "Serveur", "le classement fait à l'écran n'est pas écrasé"

    async def test_la_colonne_etat_unique_vaut_constat(self, session: AsyncSession) -> None:
        """« État constaté : Cassé » pose un constat, comme une croix dans le fichier comptable."""
        acteur = await _acteur(session, "admin.imptype3@afgbank.ml")

        rapport = await importer_classeur(
            session,
            _classeur_type([
                {"Code immo": "INV00502", "Désignation": "Écran", "État constaté": "Cassé"}
            ]),
            acteur,
        )

        assert rapport["constats_enregistres"] == 1
        e = await repo.par_code_immo(session, "INV00502")
        assert e is not None
        assert e["etat_constate"] == "CASSE"


def test_le_modele_d_import_se_relit_par_l_import() -> None:
    """Le modèle fourni au métier est lu sans accroc par l'import : ses en-têtes sont reconnues."""
    from dsi360.infrastructure.ingestion_equipements import analyser_classeur
    from dsi360.infrastructure.modele_import_equipements import construire_modele

    classeur = openpyxl.load_workbook(BytesIO(construire_modele()))
    feuille = classeur["Inventaire"]
    feuille.append(
        ["INV00999", "Portable modèle", "Ordinateur portable", "SN-1", "Latitude",
         "Siège", "SI", "MAT-1", 25, "12/03/2024", 4, 850000, "Bon"]
    )
    tampon = BytesIO()
    classeur.save(tampon)

    lignes = analyser_classeur(tampon.getvalue())
    assert len(lignes) == 1
    assert lignes[0]["code_immo"] == "INV00999"
    assert lignes[0]["type"] == "Ordinateur portable"
    assert lignes[0]["etat_constate"] == "BON"

    async def test_un_rattachement_manuel_n_est_pas_defait(self, session: AsyncSession) -> None:
        """Si la DSI a désigné le détenteur, un import muet ne doit pas l'effacer."""
        acteur = await _acteur(session, "admin.imp8@afgbank.ml")
        agent = await creer_utilisateur(session, email="manuel.imp@afgbank.ml")
        await importer_classeur(session, _classeur([_ligne()]), acteur)
        cree = await repo.par_code_immo(session, "INF00208")
        assert cree is not None
        await repo.maj(session, cree["id"], {"detenteur_id": agent})
        await session.commit()

        await importer_classeur(session, _classeur([_ligne()]), acteur)

        apres = await repo.par_code_immo(session, "INF00208")
        assert apres is not None
        assert apres["detenteur_id"] == agent


class TestLignesDouteuses:
    async def test_une_ligne_sans_code_immo_est_ignoree(self, session: AsyncSession) -> None:
        """Sans clé, impossible de la reconnaître au prochain import : elle ferait doublon."""
        acteur = await _acteur(session, "admin.imp9@afgbank.ml")

        rapport = await importer_classeur(
            session, _classeur([_ligne(**{"CODE IMMO": None})]), acteur
        )

        assert rapport == {
            "total": 1,
            "crees": 0,
            "mis_a_jour": 0,
            "ignores": 1,
            "detenteurs_non_rapproches": 0,
            "avec_etat_constate": 0,
            "constats_enregistres": 0,
            "series_en_double": 0,
        }

    async def test_une_ligne_sans_designation_n_est_pas_lue(self, session: AsyncSession) -> None:
        acteur = await _acteur(session, "admin.imp10@afgbank.ml")

        rapport = await importer_classeur(
            session, _classeur([_ligne(Designtion=None), _ligne()]), acteur
        )

        assert rapport["total"] == 1, "seule la ligne exploitable est retenue"

    @pytest.mark.parametrize("faux", ["None", "N/A", "-", "  "])
    async def test_les_fausses_valeurs_ne_deviennent_pas_des_donnees(
        self, session: AsyncSession, faux: str
    ) -> None:
        """« None » n'est ni un emplacement, ni un matricule — même piège que les tickets."""
        acteur = await _acteur(session, f"admin.faux{len(faux)}@afgbank.ml")

        await importer_classeur(
            session,
            _classeur([_ligne(EMPLACEMENT=faux, MATRICULE=faux, MODELE=faux)]),
            acteur,
        )

        e = await repo.par_code_immo(session, "INF00208")
        assert e is not None
        assert e["emplacement"] is None
        assert e["matricule_brut"] is None
        assert e["modele"] is None

    async def test_un_etat_constate_est_compte_sans_etre_applique(
        self, session: AsyncSession
    ) -> None:
        """L'état d'un matériel se consigne dans une campagne, pas comme attribut permanent."""
        acteur = await _acteur(session, "admin.imp11@afgbank.ml")

        rapport = await importer_classeur(
            session, _classeur([_ligne(**{"Rebut": "X"})]), acteur
        )

        assert rapport["avec_etat_constate"] == 1
        e = await repo.par_code_immo(session, "INF00208")
        assert e is not None
        assert e["actif"] is True, "le rebut ne sort pas le matériel du parc sans campagne"

    async def test_un_fichier_qui_n_est_pas_un_inventaire_est_refuse(
        self, session: AsyncSession
    ) -> None:
        """Mieux vaut dire pourquoi que d'afficher « 0 importé »."""
        acteur = await _acteur(session, "admin.imp12@afgbank.ml")
        classeur = openpyxl.Workbook()
        feuille = classeur.active
        assert feuille is not None
        feuille.append(["Colonne A", "Colonne B"])
        feuille.append(["x", "y"])
        tampon = BytesIO()
        classeur.save(tampon)

        with pytest.raises(ValueError, match="En-têtes introuvables"):
            await importer_classeur(session, tampon.getvalue(), acteur)
