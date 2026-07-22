"""Lecture des classeurs déposés : ce que les vrais exports ont de tordu.

Ces tests naissent d'une panne observée en production : le rapport quotidien du ServiceDesk était
refusé (« fichier non reconnu ») alors qu'il était parfaitement valide. Deux travers d'export en
sont la cause, et ils reviendront — d'où ces garde-fous.
"""

import re
import zipfile
from io import BytesIO
from typing import Any

import openpyxl
import pytest

from dsi360.infrastructure import ingestion as tickets
from dsi360.infrastructure import ingestion_equipements as equipements
from dsi360.infrastructure.classeur import apercu_intitules, feuille_avec_entetes

ENTETES_TICKETS = [
    "Type d’enregistrement de service",  # apostrophe courbe, comme le fichier réel
    None,  # colonne vide au milieu de la ligne d'en-têtes
    "Statut",
    "#",
    "Catégorie",
    "Sous-catégorie",
    "Titre",
    "Demandeur",
    "Gestionnaire de processus",
    "Priorité",
    "Date de la demande",
    "Date de fermeture",
    "Time to Repair",
    "Time to Respond",
]

LIGNE_TICKET = [
    "Incident", None, "Closed", "43899", "SOFTWARE", "Windows/office", "Licence E5 user",
    "Abdoulaye DOUCARA", "FATIMETOU ZAHRA MINT SALECK", "P3", "13-01-2026 16:42",
    "02-02-2026 19:35", "482:53", "42:12",
]


def _sous_declarer_dimension(contenu: bytes) -> bytes:
    """Réécrit la feuille en annonçant une seule cellule — ce que fait l'export ServiceDesk.

    Le classeur porte `<dimension ref="A1"/>` alors qu'il compte des milliers de lignes.
    """
    source = zipfile.ZipFile(BytesIO(contenu))
    sortie = BytesIO()
    with zipfile.ZipFile(sortie, "w") as cible:
        for item in source.infolist():
            donnees = source.read(item.filename)
            if item.filename.endswith("sheet1.xml"):
                donnees = re.sub(rb'<dimension ref="[^"]*"\s*/>', b'<dimension ref="A1"/>', donnees)
            cible.writestr(item, donnees)
    return sortie.getvalue()


def _classeur(lignes: list[list[Any]], feuille_de_garde: bool = False) -> bytes:
    wb = openpyxl.Workbook()
    premiere = wb.active
    assert premiere is not None
    if feuille_de_garde:
        premiere.title = "Filtres"
        premiere.append(["Rapport ServiceDesk — page de garde"])
        feuille = wb.create_sheet("Report")
    else:
        feuille = premiere
    feuille.append(["Filtre de base : "])
    feuille.append([None, "Compagnie égal à AFG Bank Mali"])
    feuille.append([None, "Date de la demande entre 01-01-2023 - 31-12-2026"])
    for ligne in lignes:
        feuille.append(ligne)
    tampon = BytesIO()
    wb.save(tampon)
    return tampon.getvalue()


def test_une_dimension_sous_declaree_ne_tronque_pas_la_lecture() -> None:
    """`<dimension ref="A1"/>` : le classeur ment sur sa taille, on lit quand même tout.

    C'était LA panne : en lecture `read_only`, openpyxl fait confiance à cette déclaration et
    ne rend que la première ligne — le préambule. Les en-têtes devenaient invisibles et un
    rapport valide était refusé. La lecture normale recalcule l'étendue réelle.
    """
    contenu = _sous_declarer_dimension(_classeur([ENTETES_TICKETS, LIGNE_TICKET]))

    # Le piège est bien reproduit : en read_only, une seule ligne remonte.
    tronque = openpyxl.load_workbook(BytesIO(contenu), read_only=True, data_only=True)
    assert len(list(tronque.worksheets[0].iter_rows(values_only=True))) == 1

    lus = tickets.analyser_classeur(contenu)

    assert len(lus) == 1
    assert lus[0]["source_id"] == "43899"
    assert lus[0]["statut"] == "Clôturé"


def test_les_donnees_peuvent_vivre_sur_une_autre_feuille() -> None:
    """Page de garde devant, données derrière : chercher partout plutôt que refuser."""
    contenu = _classeur([ENTETES_TICKETS, LIGNE_TICKET], feuille_de_garde=True)

    lus = tickets.analyser_classeur(contenu)

    assert len(lus) == 1
    assert lus[0]["titre"] == "Licence E5 user"


def test_un_classeur_etranger_dit_ce_qu_il_contient() -> None:
    """Un refus muet laisse sans recours : on rend les intitulés lus, pour se corriger."""
    contenu = _classeur([["Colonne inconnue", "Autre colonne"], ["valeur", "valeur"]])
    classeur = openpyxl.load_workbook(BytesIO(contenu), data_only=True)

    with pytest.raises(ValueError):
        feuille_avec_entetes(classeur, tickets.ENTETES, tickets.REPERES)

    apercu = apercu_intitules(classeur)
    assert "Colonne inconnue" in apercu
    assert "Filtre de base" in apercu


def test_l_inventaire_se_lit_aussi_malgre_une_dimension_menteuse() -> None:
    """Le même travers d'export guette l'inventaire : mêmes briques, même garantie."""
    entetes = ["CODE IMMO", "Designtion", "EMPLACEMENT", "Taux", "DA", "Durée", "VA"]
    ligne = ["INF00208", "GAB Total", "GAB EXT", 25, "22-07-2005", "4 ans", 23074595]
    contenu = _sous_declarer_dimension(_classeur([entetes, ligne]))

    lus = equipements.analyser_classeur(contenu)

    assert len(lus) == 1
    assert lus[0]["code_immo"] == "INF00208"
    assert lus[0]["valeur_acquisition"] == 23074595
